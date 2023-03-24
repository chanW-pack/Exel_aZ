import os
import openpyxl

# 엑셀 파일들이 있는 디렉토리 경로
directory = './files'

# 시트 이름
sheet_name = 'DISK_SUMM'

# 결과를 저장할 리스트
result_list = []

# 디렉토리 내의 모든 파일에 대해서 반복
for filename in os.listdir(directory):
    # 파일 경로
    file_path = os.path.join(directory, filename)
    
    # 엑셀 파일인 경우
    if file_path.endswith('.xlsx'):
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(file_path)

        # 시트 이름으로 시트 가져오기
        sheet = workbook[sheet_name]

        # 시트의 모든 행에 대해서 반복하며 데이터 추출
        for row in sheet.iter_rows():
            # row에 대한 처리
            result_list.append(row)

        # 엑셀 파일 닫기
        workbook.close()

# 결과를 저장할 엑셀 파일 경로
output_file = 'output.xlsx'

# 새로운 엑셀 파일 생성
workbook = openpyxl.Workbook()

# 새로운 시트 추가 (기존 시트 복사)
workbook.copy_worksheet(workbook.active)
sheet = workbook.active
sheet.title = sheet_name

# 결과 데이터를 새로운 시트에 저장
for row in result_list:
    sheet.append(row)

# 새로운 엑셀 파일 저장
workbook.save(output_file)
