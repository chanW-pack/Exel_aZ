
import os
from time import sleep
import openpyxl as op 
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, GradientFill, Alignment, Border, Side
from tkinter import filedialog
from tkinter.filedialog import askopenfilenames
from time import sleep
import datetime



global Az_val_1, Az_val_2, add_inq, Title, hide_font

Az_val_1 = Font(name='맑은 고딕', bold=True, size=24)
Az_val_2 = Font(name='맑은 고딕', size=24)
Title = Font(name='맑은 고딕', bold=True, size=20)
add_inq = Font(name='맑은 고딕', bold=True, size=11)
hide_font = Font(color='FFFFFF')

# cell 테두리 변수
global cell_box
cell_box = Border(
    left=Side(border_style="medium", color='00000000'),
    right=Side(border_style='medium', color='00000000'),
    bottom=Side(border_style='medium', color='00000000'),
    top=Side(border_style='medium', color='00000000')
)

# 현재 날짜 변수
global d_today
d_today = datetime.date.today()


def cpu_fx():
    path = os.getcwd()
    print(path)
    wb = op.load_workbook("{}/cw_test/CPU_SUM.xlsx".format(os.getcwd()))
    ws = wb["Sheet1"]

    # cell 병합
    ws.merge_cells("B2:I3")
    ws.merge_cells("B4:I5")

    ws.merge_cells("J2:Q5")
    ws.merge_cells("R2:U5")

    # font 지정
    ws['B2'].font = Az_val_1
    ws['B4'].font = Az_val_2
    ws['J2'].font = Title
    ws['R2'].font = add_inq
           
    # cell 배경색 추가
    ws['B2'].fill = PatternFill(start_color='B4C6E7', fill_type = 'solid')
    ws['J2'].fill = PatternFill(start_color='FFE699', fill_type = 'solid')
    ws['R2'].fill = PatternFill(start_color='D6DCE4', fill_type = 'solid')

    # cell 정렬 적용 // wrap_text=자동 줄바꿈
    ws['B2'].alignment = Alignment(vertical="center")
    ws['R2'].alignment = Alignment(horizontal="left", vertical="top", wrap_text = True)
    ws['J2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)


    

    ## 0406 추가, cell 변수로 지정
    column = ws['']
    for cell in column:
        if cell.value is None:
            break
    print(cell.value)

    import openpyxl

    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook('파일명.xlsx')
    # 시트 선택하기
    sheet = workbook['시트명']
    # 시트 내에서 'cw' 단어가 있는 셀 찾기
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None and 'cw' in str(cell):
                # 'cw' 단어가 포함된 셀이 위치한 행 번호 출력
                print(cell.row)





    # cell 값 적용
    ws["B2"].value = "CPU Utilization Average (%)"
    ws["B4"].value = "=AVERAGE('CPU_ALL:CPU_ALL (4)'!J59)"
    ws["J2"].value = f"CPU Utilization 월 평균 계산기                  {d_today}"
    ws["R2"].value = "문의사항                                                     TS팀 박찬우                                              tel: 010-9085-0857             chanwoo9730@naver.com"

        #이미지 파일 경로 및 파일명
    path_img = r"C:\Users\{}\Desktop\Exel_aZ".format(os.getlogin())
    number_file = "CPU.png"
    #Image 클래스의 객체 img 선언 : Image 클래스 선언시 매개변수는 이미지 파일 경로이다.
    img = Image(path_img + "/" + number_file)
    #WorkSheet의 add_image 함수 사용 : 매개변수는 각각 Image 객체, 불러올 위치(A1)
    ws.add_image(img,"B7")

    # 결과 시트 수정
    wb.move_sheet(ws, -4) 
    
    ws = wb["Sheet1"]
    ws.title = "CPU_Result"
    
    wb.save("{}/cw_test/CPU_SUM.xlsx".format(os.getcwd()))

cpu_fx()