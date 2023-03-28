import glob
import win32com.client
import os
import datetime
from time import sleep
import openpyxl as op 
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, GradientFill, Alignment, Border, Side
import tkinter
from tkinter import filedialog
from tkinter.filedialog import askopenfilenames
#from openpyxl.chart import LineChart, Reference, BarChart, AreaChart
#from openpyxl.chart.label import DataLabelList
#from openpyxl.chart.axis import DateAxis
#from openpyxl.chart.shapes import GraphicalProperties
#from openpyxl.utils import range_boundaries

 # 엑셀 총합 파일을 담을 디렉터리 생성
def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)
    sleep(1)

# UI로 경로 불러오기
def AzData_pull():
    global list_filepath_UI, path_get
    root = tkinter.Tk()
    root.withdraw()
    path = filedialog.askdirectory(parent=root, initialdir="./", title="폴더를 선택 해 주세요")                        
    print("path : ", path)

    path_get = glob.glob(path , recursive=True)
    
    list_filepath_UI = glob.glob(path + '/*.xlsx', recursive=True)
    for i in list_filepath_UI: 
        print(i)

def AzData_select():
    global list_filepath_UI_select, path_get
    root = tkinter.Tk()
    root.withdraw()
    path = askopenfilenames(parent=root, initialdir="/", title="파일을 선택하세요.",
                           filetypes=(("xlsx 파일", "*.xlsx"), ("all files", "*.*")))
    print(path)

    path_get = glob.glob(path , recursive=True)
    
    list_filepath_UI_select = glob.glob(path + '/*.xlsx', recursive=True)
    for i in list_filepath_UI: 
        print(i)


# 엑셀 총합 파일 생성
def disk_Az():
    DataList = list_filepath_UI
    
    # win32com(pywin32)를 이용해서 엑셀 어플리케이션을 연다.
    excel = win32com.client.Dispatch("Excel.Application")
    # 실제 작동하는 것을 보고 싶을 때
    #excel.Visible = True 

    # 엑셀 어플리케이션에 새로운 Workbook 추가
    wb_new = excel.Workbooks.Add() 

    # glob 모듈로 원하는 폴더 내의 모든 xlsx 파일의 경로를 리스트로 반환
    # list_filepath = glob.glob(r'C:\Users\*\Desktop\Exel_aZ\files\*.xlsx', recursive=True)
    
    # 엑셀 시트를 추출하고 새로운 엑셀에 붙여넣는 반복문
    for filepath in DataList:

        # 받아온 엑셀 파일의 경로를 이용해 엑셀 파일 열기
        wb = excel.Workbooks.Open(filepath)

        # 새로 만든 엑셀 파일에 추가
        # 추출할wb.Worksheets("추출할 시트명").Copy(Before=붙여넣을 wb.Worksheets("기준 시트명")
        wb.Worksheets("DISK_SUMM").Copy(Before=wb_new.Worksheets("Sheet1"))
  
    path = os.getcwd()
    print(path)
    # 취합한 엑셀 파일을 저장
    wb_new.SaveAs("{}/cw_test/DISK_SUM.xlsx".format(os.getcwd()))

    # 켜져있는 엑셀 및 어플리케이션 모두 종료
    excel.Quit()


def cpu_Az():
    DataList = list_filepath_UI
    
    excel = win32com.client.Dispatch("Excel.Application")
    wb_new = excel.Workbooks.Add() 
    #list_filepath = glob.glob(r'C:\Users\*\Desktop\Exel_aZ\files\*.xlsx', recursive=True)

    for filepath in DataList:

        wb = excel.Workbooks.Open(filepath)
        wb.Worksheets("CPU_ALL").Copy(Before=wb_new.Worksheets("Sheet1"))

    #wb_new.SaveAs(r"C:\Users\{}\Desktop\Exel_aZ\test\CPU_SUM.xlsx".format(os.getlogin()))
    path = os.getcwd()
    print(path)
    wb_new.SaveAs("{}/cw_test/CPU_SUM.xlsx".format(os.getcwd()))

    excel.Quit()


def mem_Az():
    DataList = list_filepath_UI
    
    excel = win32com.client.Dispatch("Excel.Application")
    wb_new = excel.Workbooks.Add() 
    #list_filepath = glob.glob(r'C:\Users\*\Desktop\Exel_aZ\files\*.xlsx', recursive=True)

    for filepath in DataList:
        
        wb = excel.Workbooks.Open(filepath)
        wb.Worksheets("MEM").Copy(Before=wb_new.Worksheets("Sheet1"))

    #wb_new.SaveAs(r"C:\Users\{}\Desktop\Exel_aZ\test\MEM_SUM.xlsx".format(os.getlogin()))
    path = os.getcwd()
    print(path)
    wb_new.SaveAs("{}/cw_test/MEM_SUM.xlsx".format(os.getcwd()))

    excel.Quit()


# font 변수
global Az_val_1, Az_val_2, add_inq, Title, hide_font

Az_val_1 = Font(name='맑은 고딕', bold=True, size=24)
Az_val_2 = Font(name='맑은 고딕', size=24)
Title = Font(name='맑은 고딕', bold=True, size=20)
add_inq = Font(name='맑은 고딕', bold=True, size=11)
hide_font = Font(color='FFFFFF')

# cell 테두리 변수
global cell_box
cell_box = Border(
    left=Side(border_style="medium", color='00000000'),
    right=Side(border_style='medium', color='00000000'),
    bottom=Side(border_style='medium', color='00000000'),
    top=Side(border_style='medium', color='00000000')
)

# 현재 날짜 변수
global d_today
d_today = datetime.date.today()


def disk_fx():
    path = os.getcwd()
    print(path)
    #wb = op.load_workbook(r"C:\Users\{}\Desktop\Exel_aZ\test\DISK_SUM.xlsx".format(os.getlogin())) #Workbook 객체 생성
    wb = op.load_workbook("{}/cw_test/DISK_SUM.xlsx".format(os.getcwd()))
    ws = wb["Sheet1"] 

    # cell 병합
    ws.merge_cells("B2:E3")
    ws.merge_cells("F2:I3")
    ws.merge_cells("B4:E5")
    ws.merge_cells("F4:I5")

    ws.merge_cells("J2:Q5")
    ws.merge_cells("R2:U5")

    # font 지정
    ws['B2'].font = Az_val_1
    ws['F2'].font = Az_val_1
    ws['B4'].font = Az_val_2
    ws['F4'].font = Az_val_2
    ws['J2'].font = Title
    ws['R2'].font = add_inq

    # cell 테두리 지정
    #ws['A1'].border = cell_box
    #ws['E1'].border = cell_box
    #ws['E10'].border = cell_box
                   
    # cell 배경색 추가
    ws['B2'].fill = PatternFill(start_color='C6E0B4', fill_type = 'solid') 
    ws['F2'].fill = PatternFill(start_color='B4C6E7', fill_type = 'solid')
    ws['J2'].fill = PatternFill(start_color='FFE699', fill_type = 'solid')
    ws['R2'].fill = PatternFill(start_color='D6DCE4', fill_type = 'solid')

    # cell 정렬 적용 // wrap_text=자동 줄바꿈
    ws['R2'].alignment = Alignment(horizontal="left", vertical="top", wrap_text = True)
    ws['J2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)

    # cell 값 적용
    
    
    ws["B2"].value = "Disk Read MB/s"
    ws["F2"].value = "Disk Write MB/s"
    ws["B4"].value = "=AVERAGE('DISK_SUMM:DISK_SUMM (4)'!B59)/1000"
    ws["F4"].value = "=AVERAGE('DISK_SUMM:DISK_SUMM (4)'!C59)/1000"
    ws["J2"].value = f"Disk Read/Write 월 평균 계산기               {d_today}"
    ws["R2"].value = "문의사항                                                     TS팀 박찬우                                              tel: 010-9085-0857             chanwoo9730@naver.com"

    #이미지 파일 경로 및 파일명
    path_img = r"C:\Users\{}\Desktop\Exel_aZ".format(os.getlogin())
    number_file = "DIsk.png"
    #Image 클래스의 객체 img 선언 : Image 클래스 선언시 매개변수는 이미지 파일 경로이다.
    img = Image(path_img + "/" + number_file)
    #WorkSheet의 add_image 함수 사용 : 매개변수는 각각 Image 객체, 불러올 위치(A1)
    ws.add_image(img,"B7")
 
    # 결과 시트 수정
    wb.move_sheet(ws, -4) 
    
    ws = wb["Sheet1"]
    ws.title = "DISK_Result"
    
    #wb.save("./test/DISK_SUM.xlsx")
    wb.save("{}/cw_test/DISK_SUM.xlsx".format(os.getcwd()))

    


def cpu_fx():
    #wb = op.load_workbook(r"C:\Users\{}\Desktop\Exel_aZ\test\CPU_SUM.xlsx".format(os.getlogin())) #Workbook 객체 생성
    path = os.getcwd()
    print(path)
    wb = op.load_workbook("{}/cw_test/CPU_SUM.xlsx".format(os.getcwd()))
    ws = wb["Sheet1"] #WorkSheet 객체 생성("무" Sheet)

    # cell 병합
    ws.merge_cells("B2:I3")
    ws.merge_cells("B4:I5")

    ws.merge_cells("J2:Q5")
    ws.merge_cells("R2:U5")

    # font 지정
    ws['B2'].font = Az_val_1
    ws['B4'].font = Az_val_2
    ws['J2'].font = Title
    ws['R2'].font = add_inq
           
    # cell 배경색 추가
    ws['B2'].fill = PatternFill(start_color='B4C6E7', fill_type = 'solid')
    ws['J2'].fill = PatternFill(start_color='FFE699', fill_type = 'solid')
    ws['R2'].fill = PatternFill(start_color='D6DCE4', fill_type = 'solid')

    # cell 정렬 적용 // wrap_text=자동 줄바꿈
    ws['B2'].alignment = Alignment(vertical="center")
    ws['R2'].alignment = Alignment(horizontal="left", vertical="top", wrap_text = True)
    ws['J2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)

    # cell 값 적용
    ws["B2"].value = "CPU Utilization Average (%)"
    ws["B4"].value = "=AVERAGE('CPU_ALL:CPU_ALL (4)'!J59)"
    ws["J2"].value = f"CPU Utilization 월 평균 계산기                  {d_today}"
    ws["R2"].value = "문의사항                                                     TS팀 박찬우                                              tel: 010-9085-0857             chanwoo9730@naver.com"

        #이미지 파일 경로 및 파일명
    path_img = r"C:\Users\{}\Desktop\Exel_aZ".format(os.getlogin())
    number_file = "CPU.png"
    #Image 클래스의 객체 img 선언 : Image 클래스 선언시 매개변수는 이미지 파일 경로이다.
    img = Image(path_img + "/" + number_file)
    #WorkSheet의 add_image 함수 사용 : 매개변수는 각각 Image 객체, 불러올 위치(A1)
    ws.add_image(img,"B7")

    # 결과 시트 수정
    wb.move_sheet(ws, -4) 
    
    ws = wb["Sheet1"]
    ws.title = "CPU_Result"
    
    wb.save("{}/cw_test/CPU_SUM.xlsx".format(os.getcwd()))
    


def mem_Calculation():
    #wb = op.load_workbook(r"C:\Users\{}\Desktop\Exel_aZ\test\MEM_SUM.xlsx".format(os.getlogin())) #Workbook 객체 생성
    path = os.getcwd()
    print(path)
    wb = op.load_workbook("{}/cw_test/MEM_SUM.xlsx".format(os.getcwd()))
    ws1 = wb["MEM"] 
    ws2 = wb["MEM (2)"] 
    ws3 = wb["MEM (3)"] 
    ws4 = wb["MEM (4)"] 

    ws1["R1"].value = "memtotal Avg"
    ws1["S1"].value = "memfree Avg"
    ws1["T1"].value = "cached Avg"
    ws1["U1"].value = "buffers Avg"
    ws1["R4"].value = "Memory in use"
    ws1["T4"].value = "Average Memory Usage (%)"
    ws1["R2"].value = "=AVERAGE(B2:B57)"
    ws1["S2"].value = "=AVERAGE(F2:F57)"
    ws1["T2"].value = "=AVERAGE(K2:K57)"
    ws1["U2"].value = "=AVERAGE(N2:N57)"
    ws1["R5"].value = "=SUM(S2:U2)"
    ws1["T5"].value = "=100-(R5/R2*100)"

    ws2["R1"].value = "memtotal Avg"
    ws2["S1"].value = "memfree Avg"
    ws2["T1"].value = "cached Avg"
    ws2["U1"].value = "buffers Avg"
    ws2["R4"].value = "Memory in use"
    ws2["T4"].value = "Average Memory Usage (%)"
    ws2["R2"].value = "=AVERAGE(B2:B57)"
    ws2["S2"].value = "=AVERAGE(F2:F57)"
    ws2["T2"].value = "=AVERAGE(K2:K57)"
    ws2["U2"].value = "=AVERAGE(N2:N57)"
    ws2["R5"].value = "=SUM(S2:U2)"
    ws2["T5"].value = "=100-(R5/R2*100)"

    ws3["R1"].value = "memtotal Avg"
    ws3["S1"].value = "memfree Avg"
    ws3["T1"].value = "cached Avg"
    ws3["U1"].value = "buffers Avg"
    ws3["R4"].value = "Memory in use"
    ws3["T4"].value = "Average Memory Usage (%)"
    ws3["R2"].value = "=AVERAGE(B2:B57)"
    ws3["S2"].value = "=AVERAGE(F2:F57)"
    ws3["T2"].value = "=AVERAGE(K2:K57)"
    ws3["U2"].value = "=AVERAGE(N2:N57)"
    ws3["R5"].value = "=SUM(S2:U2)"
    ws3["T5"].value = "=100-(R5/R2*100)"

    ws4["R1"].value = "memtotal Avg"
    ws4["S1"].value = "memfree Avg"
    ws4["T1"].value = "cached Avg"
    ws4["U1"].value = "buffers Avg"
    ws4["R4"].value = "Memory in use"
    ws4["T4"].value = "Average Memory Usage (%)"
    ws4["R2"].value = "=AVERAGE(B2:B57)"
    ws4["S2"].value = "=AVERAGE(F2:F57)"
    ws4["T2"].value = "=AVERAGE(K2:K57)"
    ws4["U2"].value = "=AVERAGE(N2:N57)"
    ws4["R5"].value = "=SUM(S2:U2)"
    ws4["T5"].value = "=100-(R5/R2*100)"

    wb.save("{}/cw_test/MEM_SUM.xlsx".format(os.getcwd()))

    

def mem_fx():
    #wb = op.load_workbook(r"C:\Users\{}\Desktop\Exel_aZ\test\MEM_SUM.xlsx".format(os.getlogin())) #Workbook 객체 생성
    path = os.getcwd()
    print(path)
    wb = op.load_workbook("{}/cw_test/MEM_SUM.xlsx".format(os.getcwd()))
    ws = wb["Sheet1"] #WorkSheet 객체 생성("무" Sheet)

    # cell 병합
    ws.merge_cells("B2:I3")
    ws.merge_cells("B4:I5")

    ws.merge_cells("J2:Q5")
    ws.merge_cells("R2:U5")

    # font 지정
    ws['B2'].font = Az_val_1
    ws['B4'].font = Az_val_2
    ws['J2'].font = Title
    ws['R2'].font = add_inq
           
    # cell 배경색 추가
    ws['B2'].fill = PatternFill(start_color='B4C6E7', fill_type = 'solid')
    ws['J2'].fill = PatternFill(start_color='FFE699', fill_type = 'solid')
    ws['R2'].fill = PatternFill(start_color='D6DCE4', fill_type = 'solid')

    # cell 정렬 적용 // wrap_text=자동 줄바꿈
    ws['B2'].alignment = Alignment(vertical="center")
    ws['R2'].alignment = Alignment(horizontal="left", vertical="top", wrap_text = True)
    ws['J2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)

    # cell 값 적용 
    ws["B2"].value = "Memory Usage Average (%)"
    ws["B4"].value = "=AVERAGE('MEM:MEM (4)'!T5)"
    ws["J2"].value = f"MEM Utilization 월 평균 계산기                  {d_today}"
    ws["R2"].value = "문의사항                                                     TS팀 박찬우                                              tel: 010-9085-0857             chanwoo9730@naver.com"

    #이미지 파일 경로 및 파일명
    path_img = r"C:\Users\{}\Desktop\Exel_aZ".format(os.getlogin())
    number_file = "MEM.png"
    #Image 클래스의 객체 img 선언 : Image 클래스 선언시 매개변수는 이미지 파일 경로이다.
    img = Image(path_img + "/" + number_file)
    #WorkSheet의 add_image 함수 사용 : 매개변수는 각각 Image 객체, 불러올 위치(A1)
    ws.add_image(img,"B7")

    # 결과 시트 수정
    wb.move_sheet(ws, -4) 
    
    ws = wb["Sheet1"]
    ws.title = "MEM_Result"
    
    #wb.save("./test/MEM_SUM.xlsx")
    wb.save("{}/cw_test/MEM_SUM.xlsx".format(os.getcwd()))