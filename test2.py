# 엑셀 파일 불러오기 및 새로운 시트 저장
import os
from openpyxl import Workbook
from openpyxl import load_workbook
 
path = './files'
 
def file_input(path): #각 파일 가져오기
    files = os.listdir(path)
    o_wb = Workbook()
    o_ws = o_wb["DISK_SUMM"]
    for file in files:
        wb = load_workbook(path + '\\' + file)
        ws = wb["DISK_SUMM"]
        data_input(ws, o_ws)
    o_wb.save(path + '\\' + 'Total.xlsx')
 
def data_input(ws, o_ws): # 각 파일별로 데이터 넣기
    for row in ws.iter_rows():
        data = []
        for cell in row:
            data.append(cell.value)
        o_ws.append(data)
        return o_ws
 
file_input(path)


#wb = op.load_workbook("./files/localhost_230102_0000.nmon.xlsx") #Workbook 객체 생성
#ws = wb["DISK_SUMM"] #"DISK"시트 객체 생성

#col_BB = wss["C1:C5"] #영어 column 만 가지고 오기
#col_B 값 출력하기

#for cell_0 in col_BB:
#    print(cell_0.value)  #가지고온 col_B의 인덱스 값 확인 = print(col_B)