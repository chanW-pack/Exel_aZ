import os
import openpyxl
from datetime import datetime


directory = os.getcwd() + '\\files\\'  # 현재 디렉토리 + 작업 폴더 추가
print(directory)
names = os.listdir(directory)  # 파일 및 폴더 전체를 리스트 형태로 반환
print(names)
print(directory + names[0])

wb = openpyxl.load_workbook(directory + names[0])
ws = wb.active

# 워크시트 이름 가져오기 - 리스트
sheetnames = wb.sheetnames
print(sheetnames)  # ['남문동', '홍제동']